from django.shortcuts import render
from django.http import HttpRequest, HttpResponse
import json
from .models import *
import jwt
from datetime import datetime
from security import Bcrypt
from backend_weather_iot.base_view import BaseView
from backend_weather_iot.settings import Pagination
from django.db import connection
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from authen.views import getUserFromToken

# Create your views here.

class DataDrive(BaseView):
    def hasUserFunction(sefl, access_token):
        user = getUserFromToken(access_token)
        return user.xem_export_du_lieu_moi_truong
    
    
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        if self.hasUserFunction(access_token) == False:
            res = json.dumps({
                "statusCode": 403,
                "message": "Do not have permission to use this function"
            })
            return HttpResponse(res, content_type='application/json', status=403)
        
        cursor = connection.cursor()
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')
        
        try:
            page = Pagination.get('CURRENT_PAGE') if request.GET.get('page') == None else int(request.GET.get('page'))
            item_in_page = Pagination.get('ITEM_IN_PAGE') if request.GET.get('iip') == None else int(request.GET.get('iip'))
            pages_in_webview = Pagination.get('PAGES_IN_WEBVIEW') if request.GET.get('piwv') == None else int(request.GET.get('piwv'))
        except Exception as error:
            res = json.dumps({
                "statusCode": 400,
                "message": f"{error}"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        totalRecords = None
        records = []
        
        if start_date == None and end_date != None:
            res = json.dumps({
                "statusCode": 400,
                "message": "Start date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date == None:
            res = json.dumps({
                "statusCode": 400,
                "message": "End date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date != None:
            try:
                year, month, day = start_date.split('T')[0].split('-')
                hour, minute, = start_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"Start date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            try:
                year, month, day = end_date.split('T')[0].split('-')
                hour, minute, = end_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"End date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            start_date = f"{start_date.split('T')[0]} {start_date.split('T')[1]}:00.000000"
            end_date = f"{end_date.split('T')[0]} {end_date.split('T')[1]}:59.999999"
            
            if datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S.%f') > datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S.%f'):
                res = json.dumps({
                    "statusCode": 400,
                    "message": "Start date not great than End date"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            cursor.execute(
                f"select * from esp32_data where sent_at >= %s and sent_at <= %s order by sent_at desc offset {(page - 1) * item_in_page} limit {item_in_page}",
                [f"{start_date}", f"{end_date}"]
            )
            records = cursor.fetchall()
            
            cursor.execute(
                "select count(*) as total from esp32_data where sent_at >= %s and sent_at <= %s",
                [f"{start_date}", f"{end_date}"]
            )
            totalRecords = cursor.fetchone()[0]
        
        if start_date == None and end_date == None:
            cursor.execute(
                f"select * from esp32_data order by sent_at desc offset {(page - 1) * item_in_page} limit {item_in_page}"
            )
            records = cursor.fetchall()
            
            cursor.execute(
                "select count(*) as total from esp32_data"
            )
            totalRecords = cursor.fetchone()[0]
        
        data = []
        index = 1
        for item in records:
            data.append({
                "STT": index,
                "nhietDo": item[1],
                "doAmKhongKhi": item[2],
                "anhSang": item[3],
                "doAmDat": item[4],
                "sentAt": item[5].strftime("%d-%m-%Y %H:%M:%S")
            })
            index += 1
        
        totalPages = totalRecords // item_in_page if totalRecords % item_in_page == 0 else totalRecords // item_in_page + 1
        start_page = page - (page % pages_in_webview if page % pages_in_webview != 0 else pages_in_webview) + 1
        end_page = start_page + pages_in_webview - 1 if start_page + pages_in_webview - 1 <= totalPages else totalPages

        res = json.dumps({
            "dataOfPage": list(data),
            "currentPage": page,
            "totalRecords": totalRecords,
            "totalPages": totalPages,
            "startPage": start_page,
            "endPage": end_page
        })
        return HttpResponse(res, content_type='application/json', status=200)


class DataExportExcel(BaseView):
    def hasUserFunction(sefl, access_token):
        user = getUserFromToken(access_token)
        return user.xem_export_du_lieu_moi_truong
    
    
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        if self.hasUserFunction(access_token) == False:
            res = json.dumps({
                "statusCode": 403,
                "message": "Do not have permission to use this function"
            })
            return HttpResponse(res, content_type='application/json', status=403)
        
        cursor = connection.cursor()
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')
        records = []
        
        if start_date == None and end_date != None:
            res = json.dumps({
                "statusCode": 400,
                "message": "Start date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date == None:
            res = json.dumps({
                "statusCode": 400,
                "message": "End date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date != None:
            try:
                year, month, day = start_date.split('T')[0].split('-')
                hour, minute, = start_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"Start date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            try:
                year, month, day = end_date.split('T')[0].split('-')
                hour, minute, = end_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"End date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            start_date = f"{start_date.split('T')[0]} {start_date.split('T')[1]}:00.000000"
            end_date = f"{end_date.split('T')[0]} {end_date.split('T')[1]}:59.999999"
            
            if datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S.%f') > datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S.%f'):
                res = json.dumps({
                    "statusCode": 400,
                    "message": "Start date not great than End date"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            cursor.execute(
                f"select * from esp32_data where sent_at >= %s  and sent_at <= %s order by sent_at desc",
                [f"{start_date}", f"{end_date}"]
            )
            records = cursor.fetchall()
        
        if start_date == None and end_date == None:
            cursor.execute(
                f"select * from esp32_data order by sent_at desc"
            )
            records = cursor.fetchall()
        
        # Tạo một tệp Excel mới
        workbook = openpyxl.Workbook()

        # Chọn trang tính cần làm việc
        sheet = workbook.active

        # Gộp ô A1, B1, C1 và đặt nội dung là "Title"
        sheet.merge_cells('A1:F1')
        if start_date == None and end_date == None:
            sheet['A1'] = 'Tất cả dữ liệu môi trường'
        
        if start_date != None and end_date != None:
            sheet['A1'] = f'Dữ liệu môi trường từ {start_date} đến {end_date}'
        
        bold_font = Font(bold=True)
        centered_alignment = Alignment(horizontal='center', vertical='center')
        
        sheet['A1'].font = Font(bold=True, size=12)
        sheet['A1'].alignment = centered_alignment

        # Bỏ qua một dòng
        sheet.append([])

        # Đặt định dạng in đậm cho các ô A3, B3, C3, D3, E3, F3 và thêm dữ liệu
        headers = ['STT', 'Thời gian đo', 'Nhiệt độ', 'Độ ẩm không khí', 'Cường độ ánh sáng', 'Độ ẩm đất']
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = centered_alignment

        # Thêm dữ liệu vào các ô A4, B4, C4, D4, E4, F4
        data = []
        i = 1
        for item in records:
            row = [
                i,
                item[5].strftime("%d-%m-%Y %H:%M:%S"),
                item[1],
                item[2],
                item[3],
                item[4]
            ]
            i = i + 1
            data.append(row)
        
        for row in data:
            for idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=data.index(row) + 4, column=idx)
                cell.value = value
                cell.alignment = centered_alignment
        
        col_dict = {
            1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F'
        }
        # Tự động đặt kích thước cột theo nội dung của ô (ngoại trừ cột D)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_dict[column]].width = adjusted_width
        
        # Tạo một BytesIO để lưu workbook vào bộ nhớ
        buff = io.BytesIO()
        workbook.save(buff)
        buff.seek(0)

        # Tạo một HTTP response với tệp Excel
        response = HttpResponse(content=buff.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=esp32_data_{round(datetime.timestamp(datetime.now()))}.xlsx'

        return response
