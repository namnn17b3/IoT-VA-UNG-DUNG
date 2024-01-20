from django.http import HttpRequest, HttpResponse
from backend_weather_iot.base_view import BaseView
from authen.views import getUserFromToken, valid
import jwt
import json
from django.db import connection
from backend_weather_iot.settings import Pagination
from authen.models import *
from security import *
from django.core.files.storage import default_storage
from backend_weather_iot.settings import AVATAR_USER_DIR
from datetime import datetime
from backend_weather_iot.settings import Pagination
from django.db import connection
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io

# Create your views here.


class CrudUser(BaseView):
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        cursor = connection.cursor()
        queryText = request.GET.get('queryText')
        
        try:
            page = Pagination.get('CURRENT_PAGE') if request.GET.get('page') == None else int(request.GET.get('page'))
            item_in_page = Pagination.get('ITEM_IN_PAGE') if request.GET.get('iip') == None else int(request.GET.get('iip'))
            pages_in_webview = Pagination.get('PAGES_IN_WEBVIEW') if request.GET.get('piwv') == None else int(request.GET.get('piwv'))
        except Exception as error:
            res = json.dumps({
                "statusCode": 400,
                "message": f"{error}"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        totalRecords = None
        records = []
        
        if queryText == None:
            cursor.execute(
                f'select id, email, username, phone, avatar, doan_benh, xem_export_lich_su_doan_benh, xem_export_du_lieu_moi_truong, bat_tat_led, bat_tat_pump\r\n\
                from users\r\n\
                where is_admin = false\r\n\
                order by id desc\r\n\
                offset {(page - 1) * item_in_page} limit {item_in_page}'
            )
            records = cursor.fetchall()
            
            cursor.execute(
                f'select count(*) as total\r\n\
                from users where is_admin = false'
            )
            totalRecords = cursor.fetchone()[0]
        else:
            cursor.execute(
                f"select id, email, username, phone, avatar, doan_benh, xem_export_lich_su_doan_benh, xem_export_du_lieu_moi_truong, bat_tat_led, bat_tat_pump\r\n\
                from users\r\n\
                where is_admin = false\r\n\
                and (lower(username) like lower(%s)\r\n\
                or lower(email) like lower(%s))\r\n\
                order by id desc\r\n\
                offset {(page - 1) * item_in_page}\r\n\
                limit {item_in_page}",
                [f'%{queryText}%', f'%{queryText}%']
            )
            records = cursor.fetchall()
            
            cursor.execute(
                f"select count(*) as total\r\n\
                from users\r\n\
                where is_admin = false\r\n\
                and (lower(username) like lower(%s)\r\n\
                or lower(email) like lower(%s))\r\n",
                [f'%{queryText}%', f'%{queryText}%']
            )
            totalRecords = cursor.fetchone()[0]
        
        data = []
        index = 1
        for item in records:
            data.append({
                "STT": index,
                "id": item[0],
                "email": item[1],
                "username": item[2],
                "phone": item[3],
                "avatar": 'static/authen/avatar/' + item[4],
                "doanBenh": item[5],
                "xemExportLichSuDoanBenh": item[6],
                "xemExportDuLieuMoiTruong": item[7],
                "batTatLed": item[8],
                "batTatPump": item[9]
            })
            index += 1

        totalPages = totalRecords // item_in_page if totalRecords % item_in_page == 0 else totalRecords // item_in_page + 1
        start_page = page - (page % pages_in_webview if page % pages_in_webview != 0 else pages_in_webview) + 1
        end_page = start_page + pages_in_webview - 1 if start_page + pages_in_webview - 1 <= totalPages else totalPages

        res = json.dumps({
            "dataOfPage": list(data),
            "currentPage": page,
            "totalRecords": totalRecords,
            "totalPages": totalPages,
            "startPage": start_page,
            "endPage": end_page
        })
        return HttpResponse(res, content_type='application/json', status=200)

    
    def post(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
         
        try:
            editDTO: dict = json.loads(request.POST.get('jsonData'))
        except Exception as error:
            res = json.dumps({
                "statusCode": 400,
                "message": f"Error: {error}"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        regexEmail = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
        regexPassword = r'^(?=.*[A-Z])(?=.*[a-z])(?=.*[@#$%^&+=!])(?=.*[0-9]).{8,}$'
        regexPhone = r'^0\d{9,9}$'
        if valid(regexEmail, editDTO.get('email')) == False:
            res = json.dumps({
                "statusCode": 400,
                "message": "Wrong Email Format!"
            })
            return HttpResponse(res, content_type='application/json', status=400)

        if valid(regexPassword, editDTO.get('password')) == False:
            res = json.dumps({
                "statusCode": 400,
                "message": "Wrong Password Format!"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        if valid(regexPhone, editDTO.get('phone')) == False:
            res = json.dumps({
                "statusCode": 400,
                "message": "Wrong Phone number Format!"
            })
            return HttpResponse(res, content_type='application/json', status=400)

        user = User.objects.filter(email=editDTO.get('email'))
        if len(user) == 1:
            res = json.dumps({
                "statusCode": 400,
                "message": "Email already exists!"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        avatar = request.FILES.get('avatar')
        avatar_name = 'default.jpg'
        if avatar != None:
            avatar_name = f'{round(datetime.now().timestamp() * 1000)}.jpg'
            avatar_path = AVATAR_USER_DIR + avatar_name
            writer = default_storage.open(avatar_path, 'wb')
            for chunk in avatar.chunks():
                writer.write(chunk)
        
        user = User(
            email=editDTO.get('email'),
            password=Bcrypt.hashpw(editDTO.get('password')),
            username=editDTO.get('username'),
            phone=editDTO.get('phone'),
            is_admin=False,
            doan_benh=True if editDTO.get('doanBenh') == 1 else False,
            xem_export_lich_su_doan_benh=True if editDTO.get('xemExportLichSuDoanBenh') == 1 else False,
            xem_export_du_lieu_moi_truong=True if editDTO.get('xemExportDuLieuMoiTruong') == 1 else False,
            bat_tat_led=True if editDTO.get('batTatLed') == 1 else False,
            bat_tat_pump=True if editDTO.get('batTatPump') == 1 else False,
            avatar = avatar_name
        )
        user.save()
        res = json.dumps({
            "statusCode": 201,
            "message": "Create user successfully"
        })
        return HttpResponse(res, content_type='application/json', status=201)
    
    
    def delete(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        editDTO: dict = json.loads(request.body)
        userId = editDTO.get('userId')
        user = User.objects.get(pk=editDTO.get('id'))
        if user == None:
            res = json.dumps({
                "statusCode": 400,
                "message": "User is not exists!"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        user.delete()
        res = json.dumps({
            "statusCode": 200,
            "message": "Delete user successfully"
        })
        return HttpResponse(res, content_type='application/json', status=200)


class UpdateUser(BaseView):
    def post(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)

        try:
            editDTO: dict = json.loads(request.POST.get('jsonData'))
        except Exception as error:
            res = json.dumps({
                "statusCode": 400,
                "message": f"Error: {error}"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        regexEmail = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
        regexPassword = r'^(?=.*[A-Z])(?=.*[a-z])(?=.*[@#$%^&+=!])(?=.*[0-9]).{8,}$'
        regexPhone = r'^0\d{9,9}$'
        if valid(regexEmail, editDTO.get('email')) == False:
            res = json.dumps({
                "statusCode": 400,
                "message": "Wrong Email Format!"
            })
            return HttpResponse(res, content_type='application/json', status=400)

        if valid(regexPassword, editDTO.get('password')) == False and editDTO.get('password') != '':
            res = json.dumps({
                "statusCode": 400,
                "message": "Wrong Password Format!"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        if valid(regexPhone, editDTO.get('phone')) == False:
            res = json.dumps({
                "statusCode": 400,
                "message": "Wrong Phone number Format!"
            })
            return HttpResponse(res, content_type='application/json', status=400)

        user = User.objects.get(pk=editDTO.get('id'))
        if user == None:
            res = json.dumps({
                "statusCode": 400,
                "message": "User is not exists!"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        if len(User.objects.filter(email=editDTO.get('email'))) == 1 and editDTO.get('email') != user.email:
            res = json.dumps({
                "statusCode": 400,
                "message": "Email already exists!"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        avatar = request.FILES.get('avatar')
        avatar_name = 'default.jpg'
        if avatar != None:
            avatar_name = f'{round(datetime.now().timestamp() * 1000)}.jpg'
            avatar_path = AVATAR_USER_DIR + avatar_name
            writer = default_storage.open(avatar_path, 'wb')
            for chunk in avatar.chunks():
                writer.write(chunk)
        
        user.email = editDTO.get('email')
        user.password = Bcrypt.hashpw(editDTO.get('password')) if editDTO.get('password') != '' else user.password
        user.username = editDTO.get('username')
        user.phone = editDTO.get('phone')
        user.is_admin = False
        user.doan_benh = True if editDTO.get('doanBenh') == 1 else False
        user.xem_export_lich_su_doan_benh = True if editDTO.get('xemExportLichSuDoanBenh') == 1 else False
        user.xem_export_du_lieu_moi_truong = True if editDTO.get('xemExportDuLieuMoiTruong') == 1 else False
        user.bat_tat_led = True if editDTO.get('batTatLed') == 1 else False
        user.bat_tat_pump = True if editDTO.get('batTatPump') == 1 else False
        user.avatar = avatar_name if avatar != None else user.avatar
        
        user.save()
        res = json.dumps({
            "statusCode": 200,
            "message": "Update user successfully"
        })
        return HttpResponse(res, content_type='application/json', status=200)


class ExportExcelUserData(BaseView):
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        cursor = connection.cursor()
        try:
            page = Pagination.get('CURRENT_PAGE') if request.GET.get('page') == None else int(request.GET.get('page'))
            item_in_page = Pagination.get('ITEM_IN_PAGE') if request.GET.get('iip') == None else int(request.GET.get('iip'))
            pages_in_webview = Pagination.get('PAGES_IN_WEBVIEW') if request.GET.get('piwv') == None else int(request.GET.get('piwv'))
        except Exception as error:
            res = json.dumps({
                "statusCode": 400,
                "message": f"{error}"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        records = []
        cursor.execute(
            f'select id, email, username, phone, avatar, doan_benh, xem_export_lich_su_doan_benh, xem_export_du_lieu_moi_truong, bat_tat_led, bat_tat_pump\r\n\
            from users\r\n\
            where is_admin = false\r\n'
        )
        records = cursor.fetchall()
        
        # Tạo một tệp Excel mới
        workbook = openpyxl.Workbook()

        # Chọn trang tính cần làm việc
        sheet = workbook.active
        
        sheet.column_dimensions['F'].width = 20

        # Gộp ô A1, B1, C1 và đặt nội dung là "Title"
        sheet.merge_cells('A1:K1')
        sheet['A1'] = 'Danh sách người dùng hệ thống'
        
        bold_font = Font(bold=True)
        centered_alignment = Alignment(horizontal='center', vertical='center')
        
        sheet['A1'].font = Font(bold=True, size=12)
        sheet['A1'].alignment = centered_alignment

        # Bỏ qua một dòng
        sheet.append([])

        # Đặt định dạng in đậm cho các ô A3, B3, C3, D3, E3, F3 và thêm dữ liệu
        headers = [
            'STT',
            'ID',
            'Email',
            'User name',
            'Phone',
            'Avatar',
            'Đoán bệnh của cây',
            'Xem và export lịch sử đoán bệnh của cây',
            'Xem và export dữ liệu môi trường',
            'Bật tắt LED',
            'Bật tắt máy bơm'
        ]
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = centered_alignment

        # Thêm dữ liệu vào các ô A4, B4, C4, D4, E4, F4
        data = []
        i = 1
        for item in records:
            row = [
                i,
                item[0],
                item[1],
                item[2],
                item[3],               
                f'authen/static/authen/avatar/{item[4]}',
                item[5],
                item[6],
                item[7],
                item[8],
                item[9]
            ]
            i = i + 1
            data.append(row)
        
        
        for r_idx, row in enumerate(data, start=0):
            sheet.row_dimensions[r_idx + 4].height = 120
            for idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=data.index(row) + 4, column=idx)
                if idx != 6:
                    cell.value = value
                    cell.alignment = centered_alignment
                else:
                    # Xác định vị trí của hình ảnh
                    image_path = value
                    img = Image.open(image_path)
                    
                    # img_width, img_height = img.size
                    img_width, img_height = (150, 150)

                    # Tạo một đối tượng hình ảnh trong tệp Excel
                    excel_image = ExcelImage(img)
                    excel_image.width = img_width
                    excel_image.height = img_height
                    cell.value = ''

                    # Chèn hình ảnh vào ô
                    cell.alignment = centered_alignment
                    sheet.add_image(excel_image, cell.coordinate)

        col_dict = {
            1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K'
        }
        # Tự động đặt kích thước cột theo nội dung của ô (ngoại trừ cột F)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            if column != 6:  # Bỏ qua cột F
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[col_dict[column]].width = adjusted_width

        # Tạo một BytesIO để lưu workbook vào bộ nhớ
        buff = io.BytesIO()
        workbook.save(buff)
        buff.seek(0)

        # Tạo một HTTP response với tệp Excel
        response = HttpResponse(content=buff.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=esp32_data_{round(datetime.timestamp(datetime.now()))}.xlsx'

        return response


class IoTMode(BaseView):
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        f = open('mode.txt', 'r')
        maintenance_mode, led_auto_mode, pump_auto_mode = map(int, f.readline().split())
        f.close()
        res = json.dumps({
            "ledAutoMode": led_auto_mode,
            "pumpAutoMode": pump_auto_mode,
            "maintenanceMode": maintenance_mode
        })
        return HttpResponse(res, content_type='application/json', status=200)
    
    
    def post(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        if user.is_admin == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Admin Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        f = open('mode.txt', 'r')
        maintenance_mode, led_auto_mode, pump_auto_mode = map(int, f.readline().split())
        f.close()
        
        requestDTO = json.loads(request.body)
        maintenance_mode = requestDTO.get('maintenanceMode') if requestDTO.get('maintenanceMode') != None else maintenance_mode
        led_auto_mode = requestDTO.get('ledAutoMode') if requestDTO.get('ledAutoMode') != None else led_auto_mode
        pump_auto_mode = requestDTO.get('pumpAutoMode') if requestDTO.get('pumpAutoMode') != None else pump_auto_mode
        
        f = open('mode.txt', 'w')
        f.write(f'{maintenance_mode} {led_auto_mode} {pump_auto_mode}')
        f.close()
        
        res = json.dumps({
            "status": 201,
            "message": "success"
        })
        return HttpResponse(res, content_type='application/json', status=201)
