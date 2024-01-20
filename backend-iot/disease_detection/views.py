from django.shortcuts import render
from django.http import HttpRequest, HttpResponse
import cv2
from backend_weather_iot.base_view import BaseView
import numpy as np
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.models import load_model
import json
import os
from tensorflow.keras.preprocessing.image import load_img
from backend_weather_iot.settings import UPLOAD_IMAGE_DISEASE_DIR
from datetime import datetime
import jwt
from .models import *
from authen.views import getUserFromToken
from backend_weather_iot.settings import Pagination
from django.db import connection
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io

model = load_model('model.h5')

# Create your views here.
class DiseaseDetection(BaseView):
    def hasUserFunction(self, access_token):
        user = getUserFromToken(access_token)
        return user.doan_benh
    
    
    def post(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        if self.hasUserFunction(access_token) == False:
            res = json.dumps({
                "statusCode": 403,
                "message": "Do not have permission to use this function"
            })
            return HttpResponse(res, content_type='application/json', status=403)
        
        # todo authentication
        image = request.FILES['image']
        filename = image.name
        print("@@ Input posted = ", filename)

        image_name = f'{round(1000 * datetime.timestamp(datetime.now()))}_{filename}'
        file_path = os.path.join(
            UPLOAD_IMAGE_DISEASE_DIR,
            image_name,
        )
        with open(file_path, 'wb') as destination:
            for chunk in image.chunks():
                destination.write(chunk)

        print("@@ Predicting class......")
        test_image = load_img(file_path, target_size=(128, 128))  # load image
        print("@@ Got Image for prediction")
        test_image = (
            img_to_array(test_image) / 255
        )  # convert image to np array and normalize
        test_image = np.expand_dims(test_image, axis=0)  # change dimention 3D to 4D

        result = model.predict(test_image)  # predict diseased plant or not
        print("@@ Raw result = ", result)

        pred = np.argmax(result, axis=1)[0]
        
        diseases = Diseases.objects.get(pk=pred)
        res = json.dumps({
            'tree': diseases.tree,
            'disease': diseases.disease,
            'treatment': diseases.treatment
        })
        
        user = getUserFromToken(access_token)
        
        HistoryPredictDisease.objects.create(
            user=user,
            diseases=diseases,
            image=image_name,
            sent_at=datetime.now()
        )
        
        return HttpResponse(res, content_type='application/json', status=200)


class HistoryPredictDiseaseView(BaseView): 
    def hasUserFunction(sefl, access_token):
        user = getUserFromToken(access_token)
        return user.xem_export_lich_su_doan_benh
    
    
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        if self.hasUserFunction(access_token) == False:
            res = json.dumps({
                "statusCode": 403,
                "message": "Do not have permission to use this function"
            })
            return HttpResponse(res, content_type='application/json', status=403)
        
        user = getUserFromToken(access_token)
        cursor = connection.cursor()
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')
        
        try:
            page = Pagination.get('CURRENT_PAGE') if request.GET.get('page') == None else int(request.GET.get('page'))
            item_in_page = Pagination.get('ITEM_IN_PAGE') if request.GET.get('iip') == None else int(request.GET.get('iip'))
            pages_in_webview = Pagination.get('PAGES_IN_WEBVIEW') if request.GET.get('piwv') == None else int(request.GET.get('piwv'))
        except Exception as error:
            res = json.dumps({
                "statusCode": 400,
                "message": f"{error}"
            })
            return HttpResponse(res, content_type='application/json', status=400)
        
        totalRecords = None
        records = []
        
        if start_date == None and end_date != None:
            res = json.dumps({
                "statusCode": 400,
                "message": "Start date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date == None:
            res = json.dumps({
                "statusCode": 400,
                "message": "End date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date != None:
            try:
                year, month, day = start_date.split('T')[0].split('-')
                hour, minute, = start_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"Start date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            try:
                year, month, day = end_date.split('T')[0].split('-')
                hour, minute, = end_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"End date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            start_date = f"{start_date.split('T')[0]} {start_date.split('T')[1]}:00.000000"
            end_date = f"{end_date.split('T')[0]} {end_date.split('T')[1]}:59.999999"
            
            if datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S.%f') > datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S.%f'):
                res = json.dumps({
                    "statusCode": 400,
                    "message": "Start date not great than End date"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            cursor.execute(
                f"select hpd.sent_at, d.tree, hpd.image, d.disease\r\n\
                from diseases d, history_predict_disease hpd\r\n\
                where hpd.user_id = %s and hpd.disease_id = d.id and sent_at >= %s and sent_at <= %s order by sent_at desc offset {(page - 1) * item_in_page} limit {item_in_page}",
                [user.id, f"{start_date}", f"{end_date}"]
            )
            records = cursor.fetchall()
            
            cursor.execute(
                f"select count(*) as total\r\n\
                from diseases d, history_predict_disease hpd\r\n\
                where hpd.user_id = %s and hpd.disease_id = d.id and sent_at >= %s and sent_at <= %s",
                [user.id, f"{start_date}", f"{end_date}"]
            )
            totalRecords = cursor.fetchone()[0]
        
        if start_date == None and end_date == None:
            cursor.execute(
                f"select hpd.sent_at, d.tree, hpd.image, d.disease\r\n\
                from diseases d, history_predict_disease hpd\r\n\
                where hpd.user_id = %s and hpd.disease_id = d.id order by sent_at desc offset {(page - 1) * item_in_page} limit {item_in_page}",
                [user.id]
            )
            records = cursor.fetchall()
            
            cursor.execute(
                f"select count(*) as total\r\n\
                from diseases d, history_predict_disease hpd\r\n\
                where hpd.user_id = %s and hpd.disease_id = d.id",
                [user.id]
            )
            totalRecords = cursor.fetchone()[0]
        
        data = []
        index = 1
        for item in records:
            data.append({
                "STT": index,
                "sentAt": item[0].strftime("%d-%m-%Y %H:%M:%S"),
                "tree": item[1],
                "image": f'static/disease_detection/upload/{item[2]}',
                "disease": item[3]
            })
            index += 1
        
        totalPages = totalRecords // item_in_page if totalRecords % item_in_page == 0 else totalRecords // item_in_page + 1
        start_page = page - (page % pages_in_webview if page % pages_in_webview != 0 else pages_in_webview) + 1
        end_page = start_page + pages_in_webview - 1 if start_page + pages_in_webview - 1 <= totalPages else totalPages

        res = json.dumps({
            "dataOfPage": list(data),
            "currentPage": page,
            "totalRecords": totalRecords,
            "totalPages": totalPages,
            "startPage": start_page,
            "endPage": end_page
        })
        return HttpResponse(res, content_type='application/json', status=200)


class HistoryPredictDiseaseDataExportExcel(BaseView):
    def hasUserFunction(sefl, access_token):
        user = getUserFromToken(access_token)
        return user.xem_export_lich_su_doan_benh
    
    def get(self, request: HttpRequest):
        if request.headers.get('Authorization') == None:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        access_token = request.headers.get('Authorization').split(' ')[1]
        if jwt.valid_token(access_token) == False:
            res = json.dumps({
                "statusCode": 401,
                "message": "Unauthorize!"
            })
            return HttpResponse(res, content_type='application/json', status=401)
        
        user = getUserFromToken(access_token)
        cursor = connection.cursor()
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')
        records = []
        
        if start_date == None and end_date != None:
            res = json.dumps({
                "statusCode": 400,
                "message": "Start date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if start_date != None and end_date == None:
            res = json.dumps({
                "statusCode": 400,
                "message": "End date is not null"
            })
            
            return HttpResponse(res, content_type='application/json', status=400)
        
        if self.hasUserFunction(access_token) == False:
            res = json.dumps({
                "statusCode": 403,
                "message": "Do not have permission to use this function"
            })
            return HttpResponse(res, content_type='application/json', status=403)
        
        if start_date != None and end_date != None:
            try:
                year, month, day = start_date.split('T')[0].split('-')
                hour, minute, = start_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"Start date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            try:
                year, month, day = end_date.split('T')[0].split('-')
                hour, minute, = end_date.split('T')[1].split(':')
                datetime(int(year), int(month), int(day), int(hour), int(minute))
            except Exception as error:
                res = json.dumps({
                    "statusCode": 400,
                    "message": f"End date param is wrong format: {error}"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            start_date = f"{start_date.split('T')[0]} {start_date.split('T')[1]}:00.000000"
            end_date = f"{end_date.split('T')[0]} {end_date.split('T')[1]}:59.999999"
            
            if datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S.%f') > datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S.%f'):
                res = json.dumps({
                    "statusCode": 400,
                    "message": "Start date not great than End date"
                })        
                return HttpResponse(res, content_type='application/json', status=400)
            
            cursor.execute(
                f"select hpd.sent_at, d.tree, hpd.image, d.disease\r\n\
                from diseases d, history_predict_disease hpd\r\n\
                where hpd.user_id = %s and hpd.disease_id = d.id and sent_at >= %s and sent_at <= %s order by sent_at desc",
                [user.id, f"{start_date}", f"{end_date}"]
            )
            records = cursor.fetchall()
        
        if start_date == None and end_date == None:
            cursor.execute(
                f"select hpd.sent_at, d.tree, hpd.image, d.disease\r\n\
                from diseases d, history_predict_disease hpd\r\n\
                where hpd.user_id = %s and hpd.disease_id = d.id order by sent_at desc",
                [user.id]
            )
            records = cursor.fetchall()
        
        # Tạo một tệp Excel mới
        workbook = openpyxl.Workbook()

        # Chọn trang tính cần làm việc
        sheet = workbook.active
        
        sheet.column_dimensions['D'].width = 20

        # Gộp ô A1, B1, C1 và đặt nội dung là "Title"
        sheet.merge_cells('A1:E1')
        if start_date == None and end_date == None:
            sheet['A1'] = 'Tất cả dữ đoán bệnh'
        
        if start_date != None and end_date != None:
            sheet['A1'] = f'Dữ liệu đoán bệnh từ {start_date} đến {end_date}'
        
        bold_font = Font(bold=True)
        centered_alignment = Alignment(horizontal='center', vertical='center')
        
        sheet['A1'].font = Font(bold=True, size=12)
        sheet['A1'].alignment = centered_alignment

        # Bỏ qua một dòng
        sheet.append([])

        # Đặt định dạng in đậm cho các ô A3, B3, C3, D3, E3, F3 và thêm dữ liệu
        headers = ['STT', 'Thời gian', 'Tên cây', 'Ảnh', 'Bệnh lý']
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = centered_alignment

        # Thêm dữ liệu vào các ô A4, B4, C4, D4, E4, F4
        data = []
        i = 1
        for item in records:
            row = [
                i,
                item[0].strftime("%d-%m-%Y %H:%M:%S"),
                item[1],
                f'disease_detection/static/disease_detection/upload/{item[2]}',
                item[3]
            ]
            i = i + 1
            data.append(row)
        
        
        for r_idx, row in enumerate(data, start=0):
            sheet.row_dimensions[r_idx + 4].height = 120
            for idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=data.index(row) + 4, column=idx)
                if idx != 4:
                    cell.value = value
                    cell.alignment = centered_alignment
                else:
                    # Xác định vị trí của hình ảnh
                    image_path = value
                    img = Image.open(image_path)
                    
                    # img_width, img_height = img.size
                    img_width, img_height = (150, 150)

                    # Tạo một đối tượng hình ảnh trong tệp Excel
                    excel_image = ExcelImage(img)
                    excel_image.width = img_width
                    excel_image.height = img_height
                    cell.value = ''

                    # Chèn hình ảnh vào ô
                    cell.alignment = centered_alignment
                    sheet.add_image(excel_image, cell.coordinate)

        col_dict = {
            1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E'
        }
        # Tự động đặt kích thước cột theo nội dung của ô (ngoại trừ cột D)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            if column != 4:  # Bỏ qua cột D
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[col_dict[column]].width = adjusted_width

        # Tạo một BytesIO để lưu workbook vào bộ nhớ
        buff = io.BytesIO()
        workbook.save(buff)
        buff.seek(0)

        # Tạo một HTTP response với tệp Excel
        response = HttpResponse(content=buff.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=esp32_data_{round(datetime.timestamp(datetime.now()))}.xlsx'

        return response
